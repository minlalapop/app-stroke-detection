from io import BytesIO

from docx import Document
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app
from app.storage.file_storage import LocalFileStorage, get_file_storage
from app.tests.test_reports import completed_unavailable_analysis, create_doctor, login


def ready_report(client: TestClient, tmp_path) -> tuple[dict[str, str], dict]:
    app.dependency_overrides[get_file_storage] = lambda: LocalFileStorage(root=tmp_path)
    admin = login(client, "admin", "admin123")
    analysis_id = completed_unavailable_analysis(client, admin)
    doctor = create_doctor(client)
    validation = client.post(
        f"/analyses/{analysis_id}/validate",
        headers=doctor,
        json={"validation_status": "VALIDATED", "comment": "Validé pour rapport."},
    )
    assert validation.status_code == 201
    report = client.post(f"/analyses/{analysis_id}/report", headers=doctor)
    assert report.status_code == 201
    return doctor, report.json()


def test_generate_and_download_pdf_docx_xlsx(client: TestClient, tmp_path) -> None:
    doctor, report = ready_report(client, tmp_path)
    exports = {}
    for export_format in ("pdf", "docx", "xlsx"):
        response = client.post(
            f"/reports/{report['id']}/exports/{export_format}", headers=doctor
        )
        assert response.status_code == 201
        exports[export_format] = response.json()
        assert response.json()["source"] == "DETERMINISTIC"
        assert len(response.json()["sha256"]) == 64

    pdf = client.get(
        f"/exports/{exports['pdf']['id']}/download", headers=doctor
    )
    assert pdf.status_code == 200
    assert pdf.content.startswith(b"%PDF")

    docx = client.get(
        f"/exports/{exports['docx']['id']}/download", headers=doctor
    )
    assert docx.status_code == 200
    document = Document(BytesIO(docx.content))
    text = "\n".join(paragraph.text for paragraph in document.paragraphs)
    assert "Rapport d’aide à l’analyse de l’AVC" in text
    assert "ne remplace pas l’avis médical" in text

    xlsx = client.get(
        f"/exports/{exports['xlsx']['id']}/download", headers=doctor
    )
    assert xlsx.status_code == 200
    workbook = load_workbook(BytesIO(xlsx.content), read_only=True)
    assert workbook.sheetnames == ["Rapport"]
    assert workbook["Rapport"]["A1"].value == "Rapport d’aide à l’analyse de l’AVC"

    listed = client.get(f"/reports/{report['id']}/exports", headers=doctor)
    assert listed.status_code == 200
    assert {item["format"] for item in listed.json()} == {"pdf", "docx", "xlsx"}


def test_export_requires_authentication(client: TestClient) -> None:
    response = client.get("/exports/00000000-0000-0000-0000-000000000000/download")
    assert response.status_code == 401
